import requests
import json
import openpyxl
from requests.auth import HTTPBasicAuth
import pandas as pd
from decouple import config
 
# Config
PC_IP = config('NUTANIX_HOST')
PC_IP_DR = config('NUTANIX_HOST')
 
USERNAME=config('NUTANIX_USER')
PASSWORD=config('NUTANIX_PASSWOORD')
 
headers = {
    "Content-Type": "application/json",
    "Accept": "application/json"
}
 
list_url = f"https://{PC_IP}:9440/api/nutanix/v3/vm_recovery_points/list"
rp_detail_url = f"https://{PC_IP}:9440/api/nutanix/v3/vm_recovery_points/"
vm_detail_url = f"https://{PC_IP}:9440/api/nutanix/v3/vms/"
vm_detail_url_dr = f"https://{PC_IP_DR}:9440/api/nutanix/v3/vms/"
 
def rp(sheet):
 
    requests.packages.urllib3.disable_warnings()
 
    sheet.append(["Recovery Point Name","RP UUID", "VM Name","Status"])
 
    offset = 0
    length = 390
 
    print("Fetching recovery points...")
 
    while True:
        payload = {
            "kind": "vm_recovery_point",
            "offset": offset,
            "length": length
        }
 
        resp = requests.post(list_url, headers=headers,auth=HTTPBasicAuth(USERNAME, PASSWORD), json=payload, verify=False)
        if resp.status_code != 200:
            print(f"Failed to fetch RP list. Status Code: {resp.status_code}")
            break
 
        data = resp.json()
        rps = data.get("entities", [])
        if not rps:
            break
 
        for rp in rps:
            rp_uuid = rp.get("metadata", {}).get("uuid")
            if not rp_uuid:
                continue
 
           
            rp_resp = requests.get(rp_detail_url + rp_uuid, headers=headers,auth=HTTPBasicAuth(USERNAME, PASSWORD), verify=False)
            if rp_resp.status_code != 200:
                print(f"Failed to fetch RP detail for {rp_uuid}")
                continue
 
            rp_data = rp_resp.json()
 
            rp_name = rp_data.get("status", {}).get("name")
 
            if not rp_name:
                rp_name="Name not assigned"
 
            vm_uuid = rp_data.get("spec", {}).get("resources", {}).get("parent_vm_reference", {}).get("uuid", "")
 
            # Fetch VM name from VM UUID
            vm_name = ""
            if vm_uuid:
                vm_resp = requests.get(vm_detail_url + vm_uuid, headers=headers, auth=HTTPBasicAuth(USERNAME, PASSWORD),verify=False)
                if vm_resp.status_code == 200:
                    vm_name = vm_resp.json().get("status", {}).get("name")
                    status="VM exists"
 
                vm_resp_dr = requests.get(vm_detail_url_dr + vm_uuid, headers=headers,auth=HTTPBasicAuth(USERNAME, PASSWORD), verify=False)
                if vm_resp_dr.status_code == 200:
                    vm_name1 = vm_resp_dr.json().get("status", {}).get("name")
                    vm_name = vm_name1
                    status="Migrated/Exist at Remote Site"
 
            if vm_name== "":
                vm_name="VM Deleted"
                status="Deleted"
 
            sheet.append([rp_name, rp_uuid, vm_name,status])
           
        offset += length
 

excel_file = "recovery_points.xlsx"
workbook = openpyxl.Workbook()
 
sheet_UK1 = workbook.create_sheet("RP_list")
 
rp(sheet_UK1)
 
# Save workbook
if('Sheet' in workbook.sheetnames) :
    sheet = workbook['Sheet']
    workbook.remove(sheet)
workbook.save(excel_file)
 
print(f"\nAll data saved to {excel_file}")
 
print("\nDo you want to continue for the deletion of RPs\n")
x=int(input("Press 1 to proceed, Press 0 to exit: "))
 
if  x==1:
 
    #Enter VM list file name with extension that contains the VM RP's to be deleted
    vm_file=input("Enter the VM list file_path - All RPs for the VMs will be deleted: ")
    df=pd.read_excel(vm_file)
    vm_list=df["VM Name"].dropna().astype(str).str.strip().tolist()
 
    df2=pd.read_excel(excel_file)
    filter=df2[df2["VM Name"].astype(str).str.strip().isin(vm_list)]
 
    #Final VM list and all RPs that will be deleted
    filter.to_excel("finalVM_RP.xlsx", index=False)
 
    rp_uuid_list=filter["RP UUID"].dropna().astype(str).tolist()
 
    print(rp_uuid_list)
 
    print("\nCaution: Deleted Recovery Points will not be recovered point(Verify the 'finalVM_RP' file before proceeding)\n")
    y=input("\nType - 'yes' to proceed : ")
 
    if y=="yes":
 
        for rp in rp_uuid_list:
 
            del_req=requests.delete(rp_detail_url + rp, headers=headers,auth=HTTPBasicAuth(USERNAME, PASSWORD), verify=False)
 
            if del_req.status_code in [200,202,204]:
                print(f"RP Deleted successfully for - {rp}")
            else:
                print(f"{rp} - Deletion failed")
    else:
        exit
 
print("\n\nOperation completed successfully\n")